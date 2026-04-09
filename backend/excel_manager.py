import os
import openpyxl
from datetime import datetime
import pytz

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

CLASSES = ["CMPN A", "CMPN B", "CMPN C"]

def get_ist_time():
    ist = pytz.timezone('Asia/Kolkata')
    return datetime.now(ist)

def init_excel_for_class(class_name):
    excel_path = os.path.join(BASE_DIR, f"{class_name}_Attendance.xlsx")
    txt_path = os.path.join(BASE_DIR, f"{class_name}.txt")
    
    if os.path.exists(excel_path):
        return  # already initialized

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = class_name
    
    # Headers
    ws.cell(row=1, column=1, value="Roll Number")
    ws.cell(row=1, column=2, value="Student Name")
    
    # Read text file
    if os.path.exists(txt_path):
        with open(txt_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            row = 2
            for line in lines:
                parts = line.strip().split('\t')
                if len(parts) >= 2:
                    roll = parts[0].strip()
                    name = parts[1].strip()
                    ws.cell(row=row, column=1, value=roll)
                    ws.cell(row=row, column=2, value=name)
                    row += 1
    
    wb.save(excel_path)

def add_attendance_column(class_name, session_info):
    excel_path = os.path.join(BASE_DIR, f"{class_name}_Attendance.xlsx")
    init_excel_for_class(class_name)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Check if a column for this specific session already exists today
    current_time_str = get_ist_time().strftime("%Y-%m-%d %H:%M")
    header_val = f"{current_time_str} ({session_info})"
    
    max_col = ws.max_column
    # Add new column
    ws.cell(row=1, column=max_col+1, value=header_val)
    
    wb.save(excel_path)
    return max_col + 1  # Return the column index

def mark_attendance(class_name, roll_number, col_idx):
    excel_path = os.path.join(BASE_DIR, f"{class_name}_Attendance.xlsx")
    if not os.path.exists(excel_path):
        return False
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Find row
    found_row = -1
    for row in range(2, ws.max_row + 1):
        cell_val = str(ws.cell(row=row, column=1).value)
        if cell_val == str(roll_number):
            found_row = row
            break
            
    if found_row != -1:
        # Check if already present to avoid overriding
        current_val = ws.cell(row=found_row, column=col_idx).value
        ws.cell(row=found_row, column=col_idx, value="P")
        wb.save(excel_path)
        return True
    return False

# Initialize for all
for c in CLASSES:
    init_excel_for_class(c)
